from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
from __future__ import unicode_literals
import os
import sys
import argparse
import torch
import numpy as np
import yaml
import time, datetime
import random, pdb
import pandas as pd
from openpyxl import load_workbook
from trainer import Trainer

def create_args():
    
    # This function prepares the variables shared across demo.py
    parser = argparse.ArgumentParser()

    # Standard Args
    parser.add_argument('--gpuid', nargs="+", type=int, default=[0],
                         help="The list of gpuid, ex:--gpuid 3 1. Negative value means cpu-only")
    parser.add_argument('--log_dir', type=str, default="outputs/out",
                         help="Save experiments results in dir for future plotting!")
    parser.add_argument('--learner_type', type=str, default='default', help="The type (filename) of learner")
    parser.add_argument('--learner_name', type=str, default='NormalNN', help="The class name of learner")
    parser.add_argument('--debug_mode', type=int, default=0, metavar='N',
                        help="activate learner specific settings for debug_mode")
    parser.add_argument('--repeat', type=int, default=1, help="Repeat the experiment N times")
    parser.add_argument('--seeds', nargs="+", type=int, default=[],
                         help="seed for each repeat round")
    parser.add_argument('--overwrite', type=int, default=0, metavar='N', help='Train regardless of whether saved model exists')

    # CL Args          
    parser.add_argument('--oracle_flag', default=False, action='store_true', help='Upper bound for oracle')
    parser.add_argument('--upper_bound_flag', default=False, action='store_true', help='Upper bound')
    parser.add_argument('--memory', type=int, default=0, help="size of memory for replay")
    parser.add_argument('--temp', type=float, default=2., dest='temp', help="temperature for distillation")
    parser.add_argument('--DW', default=False, action='store_true', help='dataset balancing')
    parser.add_argument('--prompt_param', nargs="+", type=float, default=[1, 1, 1],
                         help="e prompt pool size, e prompt length, g prompt length")
    
    # new add Args
    parser.add_argument('--adaptive_pred', default=True, action='store_false', help='Disable ataptive prediction.')
    parser.add_argument('--n_centroids', type=int, default=1,
                        help='number of clustering centers')
    parser.add_argument('--crct_epochs', type=int, default=10,
                        help='number of epochs for statistics replay')
    parser.add_argument('--ca_lr', type=float, default=0.0001,
                        help='learning rate for statistics replay')
    parser.add_argument('--ca_weight_decay', type=float, default=5e-4,
                        help='weight_decay for statistics replay')
    parser.add_argument('--ca_batch_size_ratio', type=float, default=4,
                        help='ca_batch_size=ratio*batch_size')
    parser.add_argument('--pretrained_weight', type=str, default='sup1k', help='load pretrained weight')
    

    # Config Arg
    parser.add_argument('--config', type=str, default="configs/config.yaml",
                         help="yaml experiment config input")

    return parser

def get_args(argv):
    parser=create_args()
    args = parser.parse_args(argv)
    config = yaml.load(open(args.config, 'r'), Loader=yaml.Loader)
    config.update(vars(args))
    return argparse.Namespace(**config)

# want to save everything printed to outfile
class Logger(object):
    def __init__(self, name):
        self.terminal = sys.stdout
        self.log = open(name, "a")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)  

    def flush(self):
        self.log.flush()

def _set_random(seed=1):
    # from SLCA
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

if __name__ == '__main__':
    args = get_args(sys.argv[1:])
    print(args)

    # determinstic backend
    torch.backends.cudnn.deterministic=True

    # duplicate output stream to output file
    if not os.path.exists(args.log_dir): os.makedirs(args.log_dir)
    log_out = args.log_dir + '/output.log'
    sys.stdout = Logger(log_out)

    # save args
    with open(args.log_dir + '/args.yaml', 'w') as yaml_file:
        yaml.dump(vars(args), yaml_file, default_flow_style=False)
    
    metric_keys = ['acc','time','fr']
    save_keys = ['global', 'pt']
    global_only = ['time','fr']
    avg_metrics = {}
    for mkey in metric_keys: 
        avg_metrics[mkey] = {}
        for skey in save_keys: avg_metrics[mkey][skey] = []

    # load results
    if args.overwrite:
        start_r = 0
    else:
        try:
            for mkey in metric_keys: 
                for skey in save_keys:
                    if (not (mkey in global_only)) or (skey == 'global'):
                        save_file = args.log_dir+'/results-'+mkey+'/'+skey+'.yaml'
                        if os.path.exists(save_file):
                            with open(save_file, 'r') as yaml_file:
                                yaml_result = yaml.safe_load(yaml_file)
                                avg_metrics[mkey][skey] = np.asarray(yaml_result['history'])

            # next repeat needed
            start_r = avg_metrics[metric_keys[0]][save_keys[0]].shape[-1]

            # extend if more repeats left
            if start_r < args.repeat:
                max_task = avg_metrics['acc']['global'].shape[0]
                for mkey in metric_keys: 
                    avg_metrics[mkey]['global'] = np.append(avg_metrics[mkey]['global'], np.zeros((max_task,args.repeat-start_r)), axis=-1)
                    if (not (mkey in global_only)):
                        avg_metrics[mkey]['pt'] = np.append(avg_metrics[mkey]['pt'], np.zeros((max_task,max_task,args.repeat-start_r)), axis=-1)

        except:
            start_r = 0
    
    for r in range(0, args.repeat):
        start_time = time.time()
        print('************************************')
        print('* STARTING TRIAL ' + str(r+1))
        print('************************************')

        # set random seeds
        try:
            seed = args.seeds[r]
        except:
            seed = r
        random.seed(seed)
        np.random.seed(seed)
        torch.manual_seed(seed)
        torch.cuda.manual_seed(seed)

        # set up a trainer
        trainer = Trainer(args, seed, metric_keys, save_keys, r) # dataloader, model definition, train, evaluate

        # init total run metrics storage
        max_task = trainer.max_task

        if r == 0 and start_r == 0: # assign np array for all the repeats
            for mkey in metric_keys: 
                avg_metrics[mkey]['global'] = np.zeros((max_task,args.repeat))
                if (not (mkey in global_only)):
                    avg_metrics[mkey]['pt'] = np.zeros((max_task,max_task,args.repeat))

        # train model
        avg_metrics = trainer.train(avg_metrics)  

        # evaluate model
        use_general_prompt = True
        avg_metrics = trainer.evaluate(avg_metrics, use_general_prompt=use_general_prompt)    # avg_metrics from trainer.train is overwritten

        total_time = time.time() - start_time
        total_time_str = str(datetime.timedelta(seconds=int(total_time)))
        print(f"=== Total time: {total_time_str} ===")

        # save results
        for mkey in metric_keys: 
            m_dir = args.log_dir+'/results-'+mkey+'/'
            if not os.path.exists(m_dir): os.makedirs(m_dir)
            for skey in save_keys:
                if (not (mkey in global_only)) or (skey == 'global'):
                    save_file = m_dir+skey+'.yaml'
                    result=avg_metrics[mkey][skey]
                    yaml_results = {}
                    if len(result.shape) > 2:
                        yaml_results['mean'] = result[:,:,:r+1].mean(axis=2).tolist()
                        if r>1: yaml_results['std'] = result[:,:,:r+1].std(axis=2).tolist()
                        yaml_results['history'] = result[:,:,:r+1].tolist()
                    else:
                        yaml_results['mean'] = result[:,:r+1].mean(axis=1).tolist()
                        if r>1: yaml_results['std'] = result[:,:r+1].std(axis=1).tolist()
                        yaml_results['history'] = result[:,:r+1].tolist()
                    with open(save_file, 'w') as yaml_file:
                        yaml.dump(yaml_results, yaml_file, default_flow_style=False)

        # Print the summary so far
        print('===Summary of experiment repeats:',r+1,'/',args.repeat,'===')
        for mkey in metric_keys: 
            print(mkey, ' | mean:', avg_metrics[mkey]['global'][-1,:r+1].mean(), 'std:', avg_metrics[mkey]['global'][-1,:r+1].std())

        print('FAA | mean:', avg_metrics['acc']['global'][-1,:r+1].mean(), 'std:', avg_metrics['acc']['global'][-1,:r+1].std())
        print('CAA | mean:', avg_metrics['acc']['global'].mean(0)[:r+1].mean(), 'std:', avg_metrics['acc']['global'].mean(0)[:r+1].std())
        print('FR | mean:', avg_metrics['fr']['global'][-1,:r+1].mean(), 'std:', avg_metrics['fr']['global'][-1,:r+1].std())

    
    # write configs and results into xlsx
    file_path = 'results.xlsx'
    content_dict = yaml.load(open(args.log_dir+'/args.yaml', 'r'), Loader=yaml.Loader)
    # add results
    content_dict['save_folder'] = args.log_dir.split('/')[-1]
    content_dict['FAA_m'] = avg_metrics['acc']['global'][-1,:r+1].mean()
    content_dict['FAA_s'] = avg_metrics['acc']['global'][-1,:r+1].std()
    content_dict['CAA_m'] = avg_metrics['acc']['global'].mean(0)[:r+1].mean()
    content_dict['CAA_s'] = avg_metrics['acc']['global'].mean(0)[:r+1].std()
    content_dict['FR_m'] = avg_metrics['fr']['global'][-1,:r+1].mean()
    content_dict['FR_s'] = avg_metrics['fr']['global'][-1,:r+1].std()

    # prepare the contents
    keys_list = list(content_dict.keys())
    values_list = list(content_dict.values())

    data_to_write = [keys_list, values_list]

    # Check if the file exists
    if not os.path.isfile(file_path):
        # Create the Excel file with headers and initial data
        initial_data = data_to_write

        # Create a new Excel file with headers and initial data
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            df = pd.DataFrame(initial_data)
            df.to_excel(writer, index=False, header=False)  # Write headers and data to the Excel file
        print(f"New file '{file_path}' created with initial data.")

    else:
        # File exists, so append new results to the existing file
        # Load the existing workbook
        workbook = load_workbook(filename=file_path)
        # Access the active worksheet
        worksheet = workbook.active

        # Count the number of columns containing Head Items
        num_columns = worksheet.max_column
        # Get the number of rows
        num_rows = worksheet.max_row
        # pdb.set_trace()
        if len(data_to_write[0]) == num_columns:
            data_to_write = data_to_write[1:]
            print('Skip saving heads ...')
        elif len(data_to_write[0]) > num_columns:
            worksheet.insert_cols(num_columns + 1, len(data_to_write[0]) - num_columns)
        
        # for row in data_to_write:
        #     worksheet.append(row)
        for row in data_to_write:
            for idx, cell in enumerate(row):
                if isinstance(cell, list):
                    row[idx] = str(cell)
            worksheet.append(row)

        workbook.save(filename=file_path)
        print(f"New experiment results added to '{file_path}'.")


